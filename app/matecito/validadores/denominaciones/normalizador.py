# -*- coding: utf-8 -*-
"""
matecito/validadores/denominaciones/normalizador.py — Normalización de denominaciones para el
proceso de comparación con los 6 algoritmos (MATEcito).

Este modulo NO compara: solo normaliza y clasifica. La comparación la sigue
haciendo el comparador que ya existe en matecito/validadores/. Se separa a
propósito, siguiendo el criterio del proyecto: normalización, validación y
vinculación son procesos distintos.

Qué resuelve que la clave() actual no resuelve:

1. UNICODE DECORATIVO. Los nombres de perfiles de redes sociales usan
   small-caps (U+1D00-U+1D7F), matematicas (U+1D400-U+1D7FF), fullwidth y
   letras IPA como sustituto estetico de las latinas. Esas NO tienen
   descomposicion NFKD, asi que el filtro [^A-Z0-9] las borra enteras:

       'ᴄᴀʟᴅᴇʀóɴ ꏲ'  ->  'O'      (clave() actual)
       'ᴄᴀʟᴅᴇʀóɴ ꏲ'  ->  'CALDERON'  (este modulo)

   El problema real no es la perdida: es que queda una clave corta no vacia
   que despues saca Jaro-Winkler altisimo por azar contra cualquier string
   corto. Un falso positivo silencioso es peor que un fallo ruidoso.

2. ORDEN DE TOKENS. Las redes sociales dan 'Nombre Apellido'; la base da
   'Apellido Nombre' (ConNomCompleto = Ape1+Ape2+Nom1+Nom2). Jaro-Winkler,
   Levenshtein y Damerau son sensibles a posicion, asi que sobre un match
   CORRECTO dan score bajo. clave_ordenada() ordena los tokens
   alfabeticamente para que esos 3 algoritmos vuelvan a ser utiles.

   Se guardan las dos claves, no una: la diferencia entre el score crudo y
   el ordenado es en si un diagnostico ("mismo nombre, orden invertido").
"""
import re
import unicodedata

# Tokens que no aportan a la identidad y solo inflan Dice/Jaccard por ser
# frecuentes en ambos lados. Se excluyen del conjunto de tokens, pero NO se
# borran de la clave (la clave debe seguir siendo reconstruible/auditable).
STOPWORDS = {
    'DE', 'DEL', 'LA', 'LAS', 'EL', 'LOS', 'Y', 'E', 'A', 'EN', 'CON',
    'THE', 'OF', 'AND',
}

# Indicios de persona JURIDICA. No es una lista cerrada ni pretende serlo:
# alcanza para apartar los comercios antes de buscarlos en una tabla de
# personas fisicas, donde no pueden estar. El resultado va a MOTIVO, nunca
# decide una baja por si solo.
INDICIOS_JURIDICA = {
    'SA', 'SRL', 'SAS', 'SCA', 'LTDA', 'SC', 'COOP', 'COOPERATIVA',
    'TALLER', 'BARRACA', 'CENTRO', 'CLUB', 'SHOP', 'STORE', 'DISTRIBUCIONES',
    'IMPRESIONES', 'PIZZERIA', 'MILANESERIA', 'CARNICERIA', 'PANADERIA',
    'FLORERIA', 'AUTOMOVILES', 'AUTOMOTORES', 'PROPIEDADES', 'INMOBILIARIA',
    'ESTETICA', 'PELUQUERIA', 'FOTOGRAFIA', 'CONTENIDOS', 'EMPRESA',
    'ESTUDIO', 'CONSULTORA', 'SERVICIOS', 'COMERCIAL', 'MERCADO',
    'FARMACIA', 'FERRETERIA', 'RESTAURANTE', 'BAR', 'HOTEL', 'GIMNASIO',
    'ENCUENTRO', 'OFICIAL', 'CATERING', 'CONSTRUCCIONES',
}

# Roles/rubros pegados al nombre con separador, tipicos de redes sociales:
# "Hector Araujo D'Andrea | Tatuador". El separador ya lo maneja el flujo de
# pipe-split del proyecto; esto solo cubre el caso de que llegue sin partir.
SEPARADORES_ROL = ('|', '·', '•', '—', '–')

_RE_NO_ALNUM = re.compile(r'[^A-Z0-9 ]')
_RE_ESPACIOS = re.compile(r'\s+')
_RE_EMAIL = re.compile(r'^[^@\s]+@[^@\s]+\.[a-z]{2,}$', re.IGNORECASE)

# Prefijos de nombre Unicode que corresponden a una letra latina disfrazada.
_FAMILIAS_LATINAS = (
    'LATIN', 'MATHEMATICAL', 'FULLWIDTH', 'CIRCLED', 'PARENTHESIZED',
    'SQUARED', 'MODIFIER LETTER', 'NEGATIVE CIRCLED', 'NEGATIVE SQUARED',
)


def _plegar_unicode_decorativo(texto):
    """
    Reduce a ASCII las letras latinas "disfrazadas" que NFKD no descompone.

    Estrategia: primero se intenta NFKD (resuelve acentos y la mayoria de las
    compatibilidades). Si el caracter sigue sin ser ASCII, se lee su NOMBRE
    Unicode oficial: 'LATIN LETTER SMALL CAPITAL C' termina en 'C',
    'MATHEMATICAL BOLD CAPITAL A' termina en 'A'. Ese ultimo token, cuando es
    una sola letra y la familia es latina, ES la letra base.

    Lo que no pertenece a una familia latina (silabarios Yi, Cherokee,
    emojis, dingbats) se convierte en espacio: no se puede afirmar que
    represente ninguna letra concreta, y adivinar seria inventar dato.
    """
    if not texto:
        return ''
    salida = []
    for ch in texto:
        if ch.isascii():
            salida.append(ch)
            continue
        desc = unicodedata.normalize('NFKD', ch)
        desc = ''.join(c for c in desc if not unicodedata.combining(c))
        if desc.isascii() and desc.strip():
            salida.append(desc)
            continue
        try:
            nombre = unicodedata.name(ch)
        except ValueError:
            salida.append(' ')
            continue
        partes = nombre.split()
        if (partes and len(partes[-1]) == 1 and partes[-1].isalpha()
                and any(f in nombre for f in _FAMILIAS_LATINAS)):
            salida.append(partes[-1])
        else:
            salida.append(' ')
    return ''.join(salida)


def clave(texto):
    """
    Clave canonica de comparacion. Mismo criterio que FNC_NYVD_CLAVE y que
    clave() del normalizador de domicilios (mayusculas -> sin acentos ->
    solo A-Z0-9 -> espacios colapsados, con N~ -> N porque el padron fue
    depurado con ese criterio), mas el plegado de Unicode decorativo.
    """
    if not texto:
        return ''
    t = str(texto)
    t = _plegar_unicode_decorativo(t)
    t = t.replace('ñ', 'N').replace('Ñ', 'N')
    t = unicodedata.normalize('NFKD', t)
    t = ''.join(c for c in t if not unicodedata.combining(c))
    t = t.upper()
    t = _RE_NO_ALNUM.sub(' ', t)
    return _RE_ESPACIOS.sub(' ', t).strip()


def clave_ordenada(texto_o_clave, ya_es_clave=False):
    """
    Clave con los tokens ordenados alfabeticamente. Neutraliza la diferencia
    de orden entre 'Nombre Apellido' (redes sociales) y 'Apellido Nombre'
    (ConNomCompleto), devolviendole sentido a Jaro-Winkler / Levenshtein /
    Damerau, que de otro modo penalizan un match correcto.

    Las stopwords se mantienen: quitarlas aca cambiaria la distancia de
    edicion de forma asimetrica segun cuantas tenga cada lado.
    """
    k = texto_o_clave if ya_es_clave else clave(texto_o_clave)
    if not k:
        return ''
    return ' '.join(sorted(k.split(' ')))


def tokens(texto_o_clave, largo_minimo=2, ya_es_clave=False):
    """
    Conjunto de tokens significativos, para Overlap / Dice / Jaccard y para
    el blocking por indice invertido. Excluye stopwords y tokens de 1 sola
    letra (iniciales), que generan coincidencias espurias masivas.
    """
    k = texto_o_clave if ya_es_clave else clave(texto_o_clave)
    if not k:
        return set()
    return {t for t in k.split(' ')
            if len(t) >= largo_minimo and t not in STOPWORDS}


def normalizacion_destructiva(original, k):
    """
    Detecta el caso en que la normalizacion perdio tanto texto que la clave
    resultante ya no representa al original. Sin esta guarda, un nombre
    escrito enteramente en un alfabeto no mapeable queda como una clave de
    1-2 caracteres que despues saca scores altisimos por azar.

    Devuelve True si hay que apartar el registro en vez de compararlo.
    """
    if not original:
        return False
    letras_orig = sum(1 for c in str(original) if c.isalnum())
    letras_clave = sum(1 for c in k if c.isalnum())
    if letras_orig >= 4 and letras_clave < max(3, letras_orig * 0.5):
        return True
    if letras_orig >= 4 and letras_clave == 0:
        return True
    return False


def clasificar(original, k=None):
    """
    Clasifica la denominacion antes de compararla y devuelve
    (tipo, motivo). Tipos:

      'F'      probable persona fisica
      'J'      probable persona juridica / comercio
      'RUIDO'  el campo no contiene una denominacion (email, username,
               nombre no normalizable, un solo token)

    El tipo NO decide nada por si solo: determina contra que universo se
    busca el candidato, y va a MOTIVO para que la revision manual entienda
    por que un registro no tuvo candidatos.
    """
    if k is None:
        k = clave(original)
    crudo = str(original or '').strip()

    if not crudo:
        return 'RUIDO', 'DENOMINACION_VACIA'
    if _RE_EMAIL.match(crudo):
        return 'RUIDO', 'CAMPO_NOMBRE_CONTIENE_EMAIL'
    if normalizacion_destructiva(crudo, k):
        return 'RUIDO', 'NOMBRE_NO_NORMALIZABLE_A_ASCII'
    # Username volcado en el campo nombre: sin espacios, con separadores
    # tecnicos, y sin ninguna mayuscula intermedia que sugiera nombre propio.
    if ' ' not in crudo and ('_' in crudo or '.' in crudo) and crudo.islower():
        return 'RUIDO', 'CAMPO_NOMBRE_CONTIENE_USERNAME'

    toks = tokens(k, ya_es_clave=True)
    if toks & INDICIOS_JURIDICA:
        indicio = sorted(toks & INDICIOS_JURIDICA)[0]
        return 'J', f'INDICIO_JURIDICA({indicio})'
    if len(toks) <= 1:
        return 'RUIDO', 'DENOMINACION_DE_UN_SOLO_TOKEN_NO_RESOLUBLE'
    return 'F', ''


def normalizar(original):
    """
    Punto de entrada. Devuelve el dict que alimenta una fila del lado A o B
    de la tabla resultante de comparacion.
    """
    crudo = str(original or '').strip()
    # Rol/rubro pegado con separador: se conserva solo el primer segmento
    # como denominacion, el resto es descripcion comercial.
    for sep in SEPARADORES_ROL:
        if sep in crudo:
            crudo = crudo.split(sep, 1)[0].strip()
            break
    k = clave(crudo)
    tipo, motivo = clasificar(crudo, k)
    return {
        'DENOMINACION': crudo,
        'CLAVE': k,
        'CLAVE_ORDENADA': clave_ordenada(k, ya_es_clave=True),
        'TOKENS': tokens(k, ya_es_clave=True),
        'TIPO': tipo,
        'MOTIVO_NORM': motivo,
    }


# =====================================================================
# LECTURA DEL EXCEL DE REDES SOCIALES
# =====================================================================
COLUMNAS_ESPERADAS = ('NOMBRE', 'USERNAME', 'EMAIL', 'TELEFONO')


def leer_excel_contactos(ruta, col_denominacion='NOMBRE'):
    """
    Lee el Excel de contactos de redes sociales. El encabezado NO esta en la
    primera fila (el archivo trae dos lineas de titulo y una vacia), asi que
    se autodetecta buscando la primera fila que contenga las columnas
    esperadas, en vez de hardcodear un offset que cambie entre corridas.

    Devuelve (encabezados, lista_de_dicts).
    """
    from openpyxl import load_workbook

    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.active

    encabezados = None
    filas = []
    for fila in ws.iter_rows(values_only=True):
        if encabezados is None:
            valores = {str(c).strip().upper() for c in fila if c is not None}
            if col_denominacion.upper() in valores:
                encabezados = [str(c).strip() if c is not None else '' for c in fila]
            continue
        if all(c is None or str(c).strip() == '' for c in fila):
            continue
        filas.append(dict(zip(encabezados, fila)))
    wb.close()

    if encabezados is None:
        raise ValueError(
            f"No se encontro la fila de encabezado con la columna "
            f"'{col_denominacion}' en {ruta}"
        )
    return encabezados, filas


def detectar_codificacion(ruta):
    """
    Devuelve la primera codificación que logra leer el archivo entero.

    Los CSV exportados desde Excel en español vienen casi siempre en
    CP1252/Latin-1, no en UTF-8. Abrirlos como UTF-8 rompe en la primera
    tilde o Ñ, y abrirlos como Latin-1 cuando en realidad son UTF-8
    convierte cada acento en mojibake ('JosÃ©'), que después el plegado
    Unicode NO puede deshacer. Por eso se prueba en orden de más estricto a
    más permisivo: utf-8 falla ruidosamente si el archivo no lo es, y
    latin-1 nunca falla, así que va último como red.
    """
    for enc in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            with open(ruta, 'r', encoding=enc) as f:
                f.read()
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return 'latin-1'


def detectar_delimitador(ruta, codificacion):
    """
    Sniffing acotado a los cuatro separadores que aparecen en la práctica.
    El punto y coma es el default de Excel en configuración regional
    española, así que es tan probable como la coma.
    """
    import csv
    with open(ruta, 'r', encoding=codificacion, newline='') as f:
        muestra = f.read(8192)
    try:
        return csv.Sniffer().sniff(muestra, delimiters=',;\t|').delimiter
    except Exception:
        # Sin sniffing confiable: gana el que más aparece en la muestra.
        conteos = {d: muestra.count(d) for d in (';', ',', '\t', '|')}
        return max(conteos, key=conteos.get) if any(conteos.values()) else ','


def leer_csv_contactos(ruta, col_denominacion='NOMBRE'):
    """
    Lee el CSV de contactos de redes sociales.

    Mismo criterio que leer_excel_contactos: el encabezado NO se asume en la
    primera fila, se busca la primera que contenga la columna esperada. Un
    CSV exportado del mismo origen arrastra las mismas líneas de título que
    el xlsx, y hardcodear un offset se rompe en cuanto cambia el export.

    Devuelve (encabezados, lista_de_dicts).
    """
    import csv

    codificacion = detectar_codificacion(ruta)
    delimitador = detectar_delimitador(ruta, codificacion)

    with open(ruta, 'r', encoding=codificacion, newline='') as f:
        crudas = list(csv.reader(f, delimiter=delimitador))

    encabezados = None
    filas = []
    for fila in crudas:
        if encabezados is None:
            valores = {str(c).strip().upper() for c in fila if c is not None}
            if col_denominacion.upper() in valores:
                encabezados = [str(c).strip() if c is not None else ''
                               for c in fila]
            continue
        if all(c is None or str(c).strip() == '' for c in fila):
            continue
        # Filas con menos celdas que el encabezado se rellenan en vez de
        # descartarse: un campo final vacío sin comilla es habitual y no
        # invalida la fila.
        fila = list(fila) + [''] * (len(encabezados) - len(fila))
        filas.append(dict(zip(encabezados, fila)))

    if encabezados is None:
        raise ValueError(
            f"No se encontro la fila de encabezado con la columna "
            f"'{col_denominacion}' en {ruta}. Codificacion detectada: "
            f"{codificacion}, delimitador: {delimitador!r}. "
            f"Primera fila leida: {crudas[0][:8] if crudas else '(archivo vacio)'}")
    return encabezados, filas


def leer_contactos(ruta, col_denominacion='NOMBRE'):
    """Despacha a xlsx o csv segun la extension."""
    if str(ruta).lower().endswith(('.xlsx', '.xlsm')):
        return leer_excel_contactos(ruta, col_denominacion)
    return leer_csv_contactos(ruta, col_denominacion)


def diagnostico(ruta, col_denominacion='NOMBRE'):
    """
    Reporte previo a cualquier comparacion: cuantas filas son comparables,
    cuantas son ruido y por que. Sirve para decidir el filtro antes de
    generar la tabla resultante, no despues.
    """
    encabezados, filas = leer_contactos(ruta, col_denominacion)
    print(f"Archivo: {ruta}")
    print(f"Columnas detectadas ({len(encabezados)}): {', '.join(encabezados)}")
    print(f"Filas de datos: {len(filas)}\n")

    conteo = {}
    problemas = []
    for i, f in enumerate(filas, start=1):
        n = normalizar(f.get(col_denominacion))
        etiqueta = n['TIPO'] if not n['MOTIVO_NORM'] else f"{n['TIPO']}/{n['MOTIVO_NORM'].split('(')[0]}"
        conteo[etiqueta] = conteo.get(etiqueta, 0) + 1
        if n['TIPO'] != 'F':
            problemas.append((i, f.get(col_denominacion), n['CLAVE'], etiqueta))

    print("--- Clasificacion ---")
    for etiqueta in sorted(conteo, key=lambda k: -conteo[k]):
        print(f"  {conteo[etiqueta]:>4}  {etiqueta}")

    comparables = sum(v for k, v in conteo.items() if k == 'F')
    print(f"\nComparables como persona fisica: {comparables} de {len(filas)}")

    if problemas:
        print("\n--- No comparables contra tabla de personas fisicas ---")
        for i, orig, k, etiqueta in problemas:
            print(f"  fila {i:>3} | {str(orig)[:38]:<38} | clave={k[:28]:<28} | {etiqueta}")

    # Duplicados por clave: se resuelven antes de comparar, no despues.
    vistas = {}
    for i, f in enumerate(filas, start=1):
        k = clave(f.get(col_denominacion))
        if k:
            vistas.setdefault(k, []).append(i)
    dups = {k: v for k, v in vistas.items() if len(v) > 1}
    if dups:
        print("\n--- Claves duplicadas dentro del propio archivo ---")
        for k, idxs in dups.items():
            print(f"  {k[:40]:<40} filas {idxs}")

    return filas


if __name__ == '__main__':
    import sys
    ruta = sys.argv[1] if len(sys.argv) > 1 else 'El_Telegrafo_contactos.xlsx'
    diagnostico(ruta)
